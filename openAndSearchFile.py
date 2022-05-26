# from urllib.robotparser import RobotFileParser
# from numpy import equal
from openpyxl import Workbook
import os.path
# import pandas as pd
from openpyxl import load_workbook



def createFile():
    workbook = Workbook()
    workbook.save(filePath)

# Grabbing the barcodes from the excel file
def grabFileBarcodes():
    wb = load_workbook(filePath)
    source = wb["Sheet"]
    for cell in source['1']:
        if cell.value != None:
            barcodes.append(cell.value)

def removeBarCode(barcode):
    if barcode not in barcodes:
        print("Barcode: ", barcode, " is not in the datasheet!")
        return False
    else:
        global barCodesListChanged
        barCodesListChanged = True
        barcodes.remove(barcode)
        print("Barcode: ", barcode, " has been removed!")
        return True

def lookForBarCode(barcode):
    if barcode in barcodes:
        return True
    else:
        return False

def addBarcodeToList(barcode):
    if barcode not in barcodes:
        global barCodesListChanged
        barCodesListChanged = True
        barcodes.append(barcode)
        return True
    else:
        print("Barcode: ",barcode, " already in data sheet")
        return False


def putBarcodesInFile():
    if barCodesListChanged:
        wb = load_workbook(filePath)
        ws = wb.active
        ws.delete_rows(1,2)
        ws.append(barcodes)
        wb.save(filename=filePath)
        print("File changed")
    else:
        print("file not changed")


desktop = os.path.normpath(os.path.expanduser("~/Desktop")) # Gabbing the path to the desktop
filePath = desktop + '/barcodes.xlsx'     # Adding the name of the file to the desktop path
barCodesListChanged = False               # This variable will let us know if something was changed in the list of barcodes    
barcodes = []                             # This list will hold all the barcodes we have in the spreadsheet

# Checking to see if the excel spreadsheet already exists or not, if it does not, we create it
def beginModule():
    print("started module")
    if os.path.isfile(filePath):
        grabFileBarcodes()
        print ("File exists")
    else:
        createFile()

def printBarCodes():
    print(barcodes)

# putBarcodesInFile()
# print(barcodes)



